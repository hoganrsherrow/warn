# import bs4
from bs4 import BeautifulSoup
import requests
import json
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import time
import pdfplumber
from io import BytesIO
import openpyxl
from urllib.parse import urljoin


def write_json_file(results_arr):
    output = {
        "results": results_arr
    }
    with open("results.json", "w") as json_file:
        json.dump(output, json_file)


def scrape_web_tx(url="https://www.twc.texas.gov/businesses/worker-adjustment-and-retraining-notification-warn-notices"):
    driver = webdriver.Chrome()

    # Open the URL
    driver.get(url)

    # Wait for the page to load
    wait = WebDriverWait(driver, 10)

    # Find all the links on the page
    links = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//a[contains(@href, '.xlsx')]")))

    # Download all .xlsx files
    for link in links:
        response = requests.get(link.get_attribute('href'))

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook.active
        col_headers = [cell.value for cell in sheet[1]]

        print(f"Contents of {link}:")
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_str = ""
            for header, value in zip(col_headers, row):
                row_str += f"{header}: {value}, "
            print(row_str)
        print()


    # Close the browser
    driver.quit()

def scrape_web_KY(url="https://kcc.ky.gov/Pages/News.aspx"):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36"
    }
    response = requests.get(url, headers=headers)
    base_url = "https://kcc.ky.gov"
    
    
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')

        xlsx_links = [urljoin(base_url, a['href']) for a in soup.find_all('a', href=True) if a ['href'].endswith('.xlsx')]

        for xlsx_link in xlsx_links:
            response = requests.get(xlsx_link, headers=headers)

            workbook = openpyxl.load_workbook(BytesIO(response.content))

            sheet = workbook.active
            col_headers = [cell.value for cell in sheet[1]]

            print(f"Contents of {xlsx_link}:")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_str = ""
                for header, value in zip(col_headers, row):
                    row_str += f"{header}: {value}, "
                print(row_str)
            print()
    else:
        print(f"Failed to fetch the webpage. Status code: {response.status_code}, Reason: {response.reason}")


def scrape_web_TN(url="https://www.tn.gov/workforce/general-resources/major-publications0/major-publications-redirect/reports.html"):
    print("Grabbing TN web results...")
    response = requests.get(url)
    html = response.content

    # make some soup
    soup = BeautifulSoup(html, 'html.parser')

    # find all <p> tags under id='main' div
    main_div = soup.find('div', {'id': 'main'})
    p_tags = main_div.find_all('p')

    results = []
    # iterate through the p_tags
    for p_tag in p_tags:
        results.append(p_tag.text.replace('\xa0', ' '))
    return results

def scrape_web_WA(url="https://fortress.wa.gov/esd/file/warn/Public/SearchWARN.aspx"):
    driver = webdriver.Chrome()
    driver.get(url)

    page_number = 1

    while True:  # Loop through all the pages
        time.sleep(5)

        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        body = soup.find('body')

        # table headings
        headers = []
        th_tags = body.find_all('th')
        for th in th_tags:
            headers.append(th.text)

        print(f"Headers... {headers}")

        # find text
        tr_tags = body.find_all('tr')
        for tr_tag in tr_tags:
            tr_str = ""
            td_tags = tr_tag.find_all('td')
            if len(td_tags) == len(headers):
                for i in range(len(td_tags)):
                    tr_str += headers[i]
                    tr_str += ": "
                    tr_str += td_tags[i].text.strip()
                    tr_str += " "
            tr_str += "State: WA"
            print(tr_str)

        # Find the next page number link
        page_number += 1
        try:
            next_page_link = driver.find_element("xpath", f'//a[@href="javascript:__doPostBack(\'ucPSW$gvMain\',\'Page${page_number}\')"]')
        except NoSuchElementException:
            break  # Exit the loop if there is no next page number link

        if next_page_link:
            next_page_link.click()  # Click the next page number link


def scrape_web_CA(url="https://edd.ca.gov/en/Jobs_and_Training/Layoff_Services_WARN"):
    # driver = webdriver.Chrome()
    # driver.get(url)

    # time.sleep(5)
    # html = driver.page_source
    # driver.quit()

    # soup = BeautifulSoup(html, 'html.parser')
    # pdf_link = soup.find("a", href=lambda href: href and href.endswith(".pdf"))["href"]
    pdf_link = "https://edd.ca.gov/siteassets/files/jobs_and_training/pubs/warn-report-for-7-1-2021-to-06-30-2022.pdf"
    
    # download the pdf file
    pdf_response = requests.get(pdf_link)
    pdf_file = BytesIO(pdf_response.content)

    # process the PDF using pdfplumber
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            table=page.extract_table()
            for row in table:
                print(f"Notice Date: {row[0]}, Effective Date: {row[2]}, Company: {row[3]}, Affected Workers: {row[5]}, {row[4]}")